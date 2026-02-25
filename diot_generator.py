#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DIOT 2025+ batch (.txt) generator for "Registro de Facturacion 2026vN".

Acepta el libro en .xlsb (pyxlsb), .xlsm/.xlsx/.xltm/.xltx (openpyxl).
En los formatos OOXML se leen los VALORES EN CACHE de las formulas (data_only),
que es lo que Excel guarda al salvar el archivo.

Reads the workbook READ-ONLY; writes only the .txt. Never modifies the workbook.

Output filename (matches the old macro convention):
    Normal:        "01. Ene 2026  N DIOT Declaracion.txt"     (2 spaces before N)
    Complementaria "01. Ene 2026 C1 DIOT Declaracion.txt"     (1 space before C#)
The double-space-before-N is intentional: ASCII sort puts N above C1, C2, ...

Three provider buckets, aggregated by RFC (tipo tercero 04 = nacional),
mirroring Resumen / Declaracion exactly:
  A) PUE bancarizadas deducibles  -> RecibidasXML
  B) Efectivo <= LimiteEfectivo   -> RecibidasXML  (FormaPago 01, no combustible)
  C) REP (PPD pagados en el mes)  -> PagosRecibidasXML

Per-row 54-field layout: 1=tipo tercero(04), 2=tipo op(85/03/06), 3=RFC,
12=Base 16%, 22=IVA acreditable, 48=IVA retenido (si aplica).

Rounding: Art. 20 CFF (0.01-0.50 -> abajo; 0.51-0.99 -> arriba).

Invariante SAT: el validador del SAT calcula el "IVA pagado" del renglon a
partir de la base y rechaza la linea si el IVA acreditable lo excede. No
rechaza solo cuando redondea: el acuse ErroresCargaMasiva de mayo 2026 marco
renglones donde trunc(base*0.16) < iva aunque round(base*0.16) == iva. Por eso
la base se sube hasta el minimo entero que cumple  (base*16)//100 >= iva
(aritmetica entera, sin errores de coma flotante). Esa condicion es mas
estricta que el redondeo, asi que satisface cualquiera de las dos reglas.
Costo tipico: +1 a +6 pesos por renglon (~0.01% de la base) = irrelevancia.

CLI:
    python diot_generator.py                      # abre la GUI (recomendado)
    python diot_generator.py --gui
    python diot_generator.py 6                    # Junio normal, Save As
    python diot_generator.py 6 C1                 # Junio complementaria 1
    python diot_generator.py 6 N "ruta\\out.txt"  # sin dialogos
    python diot_generator.py 6 N "out.txt" --libro "ruta\\Registro.xlsm"

Boardflare / Excel-Python:
    from diot_generator import build_diot, read_sheet
    rh, rr = read_sheet(path, "RecibidasXML")
    ph, pr = read_sheet(path, "PagosRecibidasXML")
    records, totals = build_diot(rh, rr, ph, pr, year, month, op_map={})
"""

import sys, csv, os, re, json, glob, datetime as _dt
from collections import defaultdict
from pathlib import Path

# ---------- defaults (se sobreescriben con la hoja Control si existe) ----------
USOS_DED     = {"G01", "G03"}        # Control: Usos CFDI deducibles
LIMITE_EFE   = 2000                  # Control: Limite pago en efectivo (IVA incluido)
DEFAULT_OP   = "85"                  # Otros
OPMAP_CSV    = "diot_rfc_op.csv"     # opcional: columnas  RFC,Op  (03 servicios / 06 arrend)
CONFIG_PATH  = Path.home() / ".diot_config.json"

MES_ABBR = ["", "Ene","Feb","Mar","Abr","May","Jun",
                "Jul","Ago","Sep","Oct","Nov","Dic"]
MES_LARGO = ["", "Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]


# ---------- helpers ----------
def cff_round(x):
    """Art. 20, parr. 7 CFF: 0.01-0.50 -> abajo, 0.51-0.99 -> arriba."""
    if x is None:
        return 0
    if x < 0:
        return -cff_round(-x)
    n = int(x)
    cents = int(round((x - n) * 100))
    return n + 1 if cents >= 51 else n


def ensure_invariant(base_int, iva_int):
    """Minimo entero >= base_int tal que trunc(base*0.16) >= iva_int.

    (base*16)//100 >= iva  <=>  base >= ceil(iva*100/16) = ceil(iva*25/4).
    Todo en enteros: el SAT trunca, y 1257*0.16 en float da 201.12000000000003.
    """
    if iva_int <= 0:
        return base_int
    minimo = -(-iva_int * 25 // 4)          # ceil(iva*25/4)
    return max(base_int, minimo)


def num(v):
    try: return float(v)
    except (TypeError, ValueError): return 0.0


def txt(v):
    return str(v).strip() if v is not None else ""


def is_yes(v):
    if isinstance(v, bool): return v          # fórmula cacheada como TRUE/FALSE
    return txt(v).lower() in ("si", "sí")


EXCEL_EPOCH = _dt.datetime(1899, 12, 30)   # serial 1 = 1900-01-01 (bug 1900 incluido)


def date_parts(v):
    """(año, mes) de una fecha venga como venga. (0, 0) si no se reconoce.

    .xlsb (pyxlsb) entrega las fechas como texto o como serial numérico;
    .xlsx/.xlsm (openpyxl) las entrega ya como datetime. Se aceptan las tres.
    """
    if v is None or v == "":
        return 0, 0
    if isinstance(v, _dt.datetime) or isinstance(v, _dt.date):
        return v.year, v.month
    if isinstance(v, bool):
        return 0, 0
    if isinstance(v, (int, float)):                      # serial de Excel
        try:
            d = EXCEL_EPOCH + _dt.timedelta(days=float(v))
        except (OverflowError, ValueError):
            return 0, 0
        return d.year, d.month
    s = str(v).strip()
    m = re.match(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)      # dd/mm/aaaa
    if m:
        return int(m.group(3)), int(m.group(2))
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)            # aaaa-mm-dd (ISO)
    if m:
        return int(m.group(1)), int(m.group(2))
    return 0, 0


def month_of(s):
    """'21/06/2026', '2026-06-21 12:00:00', datetime o serial de Excel -> 6"""
    return date_parts(s)[1]


def year_of(s):
    return date_parts(s)[0]


def code2(v): return txt(v)[:2]
def uso_code(v): return txt(v)[:3]


# ---------- acceso al libro (.xlsb / .xlsm / .xlsx) ----------
XLSB_EXTS     = (".xlsb",)
OPENPYXL_EXTS = (".xlsm", ".xlsx", ".xltm", ".xltx")
LIBRO_EXTS    = XLSB_EXTS + OPENPYXL_EXTS


def _grid_xlsb(path, sheet):
    import pyxlsb
    with pyxlsb.open_workbook(path) as wb:
        with wb.get_sheet(sheet) as ws:
            return [{c.c: c.v for c in r} for r in ws.rows()]


def _load_openpyxl(path, data_only):
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Para leer .xlsx/.xlsm hace falta openpyxl:  pip install openpyxl")
    return openpyxl.load_workbook(path, read_only=True, data_only=data_only,
                                  keep_links=False)


def _has_formulas(path, sheet):
    """¿La hoja tiene fórmulas? (segunda pasada, sin data_only)."""
    wb = _load_openpyxl(path, data_only=False)
    try:
        if sheet not in wb.sheetnames: return False
        for i, r in enumerate(wb[sheet].iter_rows(values_only=True)):
            if i == 0: continue                       # encabezados
            for v in r:
                if isinstance(v, str) and v.startswith("="):
                    return True
    finally:
        wb.close()
    return False


def _grid_openpyxl(path, sheet):
    """data_only=True -> valores en caché de las fórmulas (los que guardó Excel)."""
    wb = _load_openpyxl(path, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise KeyError("la hoja '%s' no existe en %s" % (sheet, os.path.basename(path)))
        grid, vivo = [], False
        for r in wb[sheet].iter_rows(values_only=True):
            d = {i: v for i, v in enumerate(r) if v is not None and v != ""}
            grid.append(d)
            if len(grid) > 1 and d:
                vivo = True
    finally:
        wb.close()
    if len(grid) > 1 and not vivo and _has_formulas(path, sheet):
        # el archivo lo escribió un programa, no Excel: las fórmulas no traen caché
        raise ValueError(
            "La hoja '%s' de %s se lee vacía: el archivo no trae los valores\n"
            "calculados de sus fórmulas. Ábrelo en Excel y guárdalo para que se\n"
            "escriba la caché, o usa la versión .xlsb."
            % (sheet, os.path.basename(path)))
    while grid and not grid[-1]:                      # filas vacías al final
        grid.pop()
    return grid


def read_grid(path, sheet):
    """Raw grid (list of {col_index: value}) — para Control y Resumen."""
    ext = os.path.splitext(str(path))[1].lower()
    if ext in OPENPYXL_EXTS:
        return _grid_openpyxl(path, sheet)
    if ext in XLSB_EXTS:
        return _grid_xlsb(path, sheet)
    raise ValueError("Formato no soportado: '%s'. Usa %s"
                     % (ext or path, ", ".join(LIBRO_EXTS)))


def read_sheet(path, sheet):
    """Returns (header_list, list_of_dict_rows)."""
    grid = read_grid(path, sheet)
    if not grid or not grid[0]: return [], []
    headers = [str(grid[0].get(i, "")).strip() for i in range(max(grid[0]) + 1)]
    hi = {h: i for i, h in enumerate(headers) if h}
    return headers, [{h: d.get(i) for h, i in hi.items()} for d in grid[1:]]


def find_col(headers, *cands):
    for c in cands:
        if c in headers: return c
    low = {h.lower(): h for h in headers}
    for c in cands:
        if c.lower() in low: return low[c.lower()]
    for c in cands:
        for h in headers:
            if c.lower() in h.lower(): return h
    raise KeyError("columna no encontrada: probé %s" % (cands,))


def find_col_opt(headers, *cands):
    """Como find_col pero devuelve None si no existe (columnas nuevas de v5)."""
    try: return find_col(headers, *cands)
    except KeyError: return None


def read_control(path):
    """Lee Control: año fiscal, límite de efectivo y usos CFDI deducibles."""
    year, limite, usos = None, None, set()
    try:
        grid = read_grid(path, "Control")
    except Exception:
        return year, limite, usos
    collecting = False
    for row in grid:
        label, val = txt(row.get(7)), row.get(8)
        if label.lower().startswith("año fiscal") and isinstance(val, (int, float)):
            year = int(val)
        elif label.lower().startswith("límite pago en efectivo") and isinstance(val, (int, float)):
            limite = float(val)
        elif label.lower().startswith("usos cfdi deducibles"):
            collecting = True
            continue
        if collecting and re.fullmatch(r"[A-Z]\d{2}", label):
            usos.add(label)
    return year, limite, usos


def read_resumen(path, month):
    """Fila del mes en Resumen -> totales esperados por bucket (para conciliar)."""
    try:
        grid = read_grid(path, "Resumen")
    except Exception:
        return None
    for row in grid:
        if isinstance(row.get(0), (int, float)) and int(row[0]) == month:
            g = lambda i: num(row.get(i))
            return {"A_n": g(14), "A_iva": g(17),
                    "B_n": g(18), "B_iva": g(22),
                    "C_n": g(23), "C_base": g(24), "C_iva": g(25)}
    return None


# ---------- core aggregation (Boardflare-callable, no I/O) ----------
def build_diot(rec_headers, rec_rows, pag_headers, pag_rows, year, month, op_map,
               limite_efe=LIMITE_EFE, usos_ded=None):
    usos_ded = usos_ded or USOS_DED

    R = lambda *c: find_col(rec_headers, *c)
    rc_fecha=R("Fecha Emision","Fecha"); rc_tipo=R("Tipo"); rc_estado=R("Estado")
    rc_metodo=R("Metodo de Pago","Metodo"); rc_banc=R("Bancarizado"); rc_uso=R("UsoCFDI")
    rc_iva=R("IVA 16%"); rc_forma=R("FormaDePago")
    rc_comb=R("Combustible"); rc_rfc=R("RFC Emisor","RFC"); rc_ret=R("Retenido IVA")
    # v5 quitó "Importe Neto"; Control dice "Límite pago en efectivo (IVA incluido)"
    # => el tope de $2,000 va contra el Total (IVA incluido).
    rc_neto=R("Total", "Importe Neto")

    P = lambda *c: find_col(pag_headers, *c)
    pc_fpago=P("FechaPago"); pc_estado=P("Estado")
    pc_banc=P("BancarizadoP","Bancarizado")
    pc_base=P("IVA 16 Base","TotalTrasladosBaseIVA16")
    pc_iva =P("IVA 16 Importe","TotalTrasladosImpuestoIVA16")
    pc_rfc =P("RFC Emisor CFDI","RFC Emisor")
    # columnas nuevas de v5 (opcionales)
    pc_rel = find_col_opt(pag_headers, "RelUUID")
    pc_dup = find_col_opt(pag_headers, "EsDuplicado")

    agg = defaultdict(lambda: {"base":0.0,"iva":0.0,"ret":0.0})
    tot = {"A_base":0.0,"A_iva":0.0,"B_base":0.0,"B_iva":0.0,
           "C_base":0.0,"C_iva":0.0,"ret":0.0,
           "A_n":0,"B_n":0,"C_n":0}
    descartados = []

    for row in rec_rows:
        rfc = row.get(rc_rfc)
        if not rfc: continue
        if month_of(row.get(rc_fecha)) != month or year_of(row.get(rc_fecha)) != year: continue
        if txt(row.get(rc_tipo)) != "Factura": continue
        # Estado es columna MANUAL (Resumen: "marcar Cancelado ..."): vacío = vigente.
        if txt(row.get(rc_estado)) == "Cancelado": continue
        if not txt(row.get(rc_metodo)).startswith("PUE"): continue
        if uso_code(row.get(rc_uso)) not in usos_ded: continue

        # base = IVA/0.16 (coincide con Declaracion C28 y absorbe el IEPS).
        iva  = num(row.get(rc_iva))
        base = iva / 0.16 if iva else 0.0
        ret  = num(row.get(rc_ret))

        if is_yes(row.get(rc_banc)):                                   # bucket A
            a = agg[rfc]; a["base"] += base; a["iva"] += iva; a["ret"] += ret
            tot["A_base"] += base; tot["A_iva"] += iva; tot["ret"] += ret; tot["A_n"] += 1
        elif (code2(row.get(rc_forma)) == "01"
              and num(row.get(rc_neto)) <= limite_efe
              and not is_yes(row.get(rc_comb))):                       # bucket B
            a = agg[rfc]; a["base"] += base; a["iva"] += iva; a["ret"] += ret
            tot["B_base"] += base; tot["B_iva"] += iva; tot["ret"] += ret; tot["B_n"] += 1

    for row in pag_rows:                                              # bucket C (REP)
        rfc = row.get(pc_rfc)
        if not rfc: continue
        if month_of(row.get(pc_fpago)) != month or year_of(row.get(pc_fpago)) != year: continue
        if txt(row.get(pc_estado)) == "Cancelado": continue
        if not is_yes(row.get(pc_banc)): continue
        if pc_dup and num(row.get(pc_dup)) == 1: continue
        if pc_rel and txt(row.get(pc_rel)) != "OK":
            # el CFDI relacionado del REP no está en RecibidasXML -> Resumen lo excluye
            descartados.append((rfc, txt(row.get(pc_rel)), num(row.get(pc_iva))))
            continue
        base = num(row.get(pc_base)); iva = num(row.get(pc_iva))
        a = agg[rfc]; a["base"] += base; a["iva"] += iva
        tot["C_base"] += base; tot["C_iva"] += iva; tot["C_n"] += 1

    records = []
    bumped = 0
    for rfc in sorted(agg):
        a = agg[rfc]
        iva_int  = cff_round(a["iva"])
        base_int = cff_round(a["base"])
        ret_int  = cff_round(a["ret"])
        if iva_int == 0 and base_int == 0 and ret_int == 0:
            continue
        adj = ensure_invariant(base_int, iva_int)                 # invariante SAT
        if adj != base_int:
            bumped += (adj - base_int)
            base_int = adj
        f = [""] * 54
        f[0]  = "04"
        f[1]  = op_map.get(rfc, DEFAULT_OP)
        f[2]  = rfc
        f[11] = str(base_int)                                     # campo 12
        f[21] = str(iva_int)                                      # campo 22
        if ret_int:
            f[47] = str(ret_int)                                  # campo 48
        records.append(f)

    tot["_bumped_pesos"] = bumped
    tot["_descartados"]  = descartados
    return records, tot


# ---------- naming / config ----------
def diot_filename(month, year, tipo):
    abbr = MES_ABBR[month]
    if tipo.upper() == "N":
        return "%02d. %s %d  N DIOT Declaración.txt" % (month, abbr, year)
    return "%02d. %s %d %s DIOT Declaración.txt" % (month, abbr, year, tipo.upper())


def detect_rfc(*paths):
    for p in paths:
        m = re.search(r'\b([A-ZÑ&]{3,4}\d{6}[A-Z\d]{3})\b', str(p or ""))
        if m: return m.group(1)
    return "RFC"


def detect_year(*paths):
    for p in paths:
        m = re.search(r'[\\/](20\d{2})[\\/]', str(p or ""))
        if m: return int(m.group(1))
    return None


def load_cfg():
    if CONFIG_PATH.exists():
        try: return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception: return {}
    return {}


def save_cfg(cfg):
    try:
        CONFIG_PATH.write_text(json.dumps(cfg, indent=2, ensure_ascii=False),
                               encoding="utf-8")
    except Exception:
        pass


def find_workbook(folder="."):
    """Busca el libro: primero por nombre, luego cualquiera. .xlsb tiene prioridad
    (trae siempre los valores calculados) y después .xlsm / .xlsx."""
    for pat in ("Registro de Facturaci*v*%s", "*%s"):
        for ext in LIBRO_EXTS:
            hits = [h for h in glob.glob(os.path.join(folder, pat % ext))
                    if not os.path.basename(h).startswith("~$")]
            if hits:
                return sorted(hits, reverse=True)[0]
    return None


def load_opmap(folder):
    op_map, path = {}, os.path.join(folder, OPMAP_CSV)
    if os.path.exists(path):
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                if r.get("RFC"):
                    op_map[r["RFC"].strip()] = str(r.get("Op", DEFAULT_OP)).strip()
    return op_map


# ---------- generación (compartida por CLI y GUI) ----------
def generate(libro, month, tipo, out_path=None, ask_path=None):
    """Devuelve (out_path, records, tot, year, rfc, resumen) o (None,...) si se cancela."""
    ext = os.path.splitext(str(libro))[1].lower()
    if ext not in LIBRO_EXTS:
        raise ValueError("Formato no soportado: '%s'.\nUsa un libro %s"
                         % (ext or libro, ", ".join(LIBRO_EXTS)))
    folder = os.path.dirname(os.path.abspath(libro))
    rfc    = detect_rfc(libro, os.getcwd())

    ctrl_year, ctrl_limite, ctrl_usos = read_control(libro)
    year   = ctrl_year or detect_year(libro, os.getcwd()) or 2026
    limite = ctrl_limite if ctrl_limite is not None else LIMITE_EFE
    usos   = ctrl_usos or USOS_DED

    rh, rr = read_sheet(libro, "RecibidasXML")
    ph, pr = read_sheet(libro, "PagosRecibidasXML")
    records, tot = build_diot(rh, rr, ph, pr, year, month, load_opmap(folder),
                              limite_efe=limite, usos_ded=usos)
    resumen = read_resumen(libro, month)

    if not out_path:
        cfg = load_cfg()
        last_dir = cfg.get(rfc, {}).get("last_save_dir") or folder
        if not os.path.isdir(last_dir): last_dir = folder
        out_path = ask_path(diot_filename(month, year, tipo), last_dir) if ask_path else None
        if not out_path:
            return None, records, tot, year, rfc, resumen
        cfg.setdefault(rfc, {})["last_save_dir"] = str(Path(out_path).parent)
        cfg.setdefault(rfc, {})["last_libro"]    = str(libro)
        save_cfg(cfg)

    with open(out_path, "w", encoding="utf-8", newline="") as fh:
        for f in records:
            fh.write("|".join(f) + "\r\n")
    return out_path, records, tot, year, rfc, resumen


def summary(out_path, records, tot, year, rfc, month, tipo, resumen, limite=LIMITE_EFE):
    base_tot = sum(int(r[11] or 0) for r in records)
    iva_tot  = sum(int(r[21] or 0) for r in records)
    ret_tot  = sum(int(r[47] or 0) for r in records if r[47])
    L = []
    L.append("=" * 72)
    L.append(" DIOT %s  -  %s %d   tipo: %s" % (rfc, MES_LARGO[month], year, tipo))
    L.append("=" * 72)
    L.append(" Proveedores en TXT:     %d" % len(records))
    L.append(" Archivo:                %s" % out_path)
    L.append("-" * 72)
    L.append(" %-24s %14s %14s   %s" % ("Bucket", "Base", "IVA", "vs Resumen"))
    rows = [("A (PUE bancarizada)", tot["A_base"], tot["A_iva"], "A_iva", "A_n", tot["A_n"]),
            ("B (efectivo <= %d)" % limite, tot["B_base"], tot["B_iva"], "B_iva", "B_n", tot["B_n"]),
            ("C (REP / PPD)",       tot["C_base"], tot["C_iva"], "C_iva", "C_n", tot["C_n"])]
    for label, b, i, key, nkey, cnt in rows:
        chk = ""
        if resumen:
            diff = i - resumen[key]
            chk = "OK" if abs(diff) < 0.01 else "DIF %.2f" % diff
            if abs(resumen[nkey] - cnt) >= 1:
                chk += " (n %d vs %d)" % (cnt, int(resumen[nkey]))
        L.append(" %-24s %14.2f %14.2f   %s" % (label, b, i, chk))
    L.append("-" * 72)
    L.append(" %-24s %14d %14d" % ("TXT TOTAL (CFF redon.)", base_tot, iva_tot))
    if ret_tot:
        L.append(" IVA retenido            %d   (verificar campo 48)" % ret_tot)
    if tot.get("_bumped_pesos"):
        L.append(" Ajuste base invariante SAT  +%d peso(s)  (trunc(base*0.16) >= IVA)"
                 % tot["_bumped_pesos"])
    for rfc_d, motivo, iva_d in tot.get("_descartados", []):
        L.append(" REP descartado: %-14s RelUUID=%-10s IVA %.2f" % (rfc_d, motivo, iva_d))
    L.append("=" * 72)
    L.append(" Compara con Declaracion C34 (VALOR ACTOS 16%) y C35 (TOTAL IVA).")
    return "\n".join(L)


# ---------- GUI ----------
def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk, scrolledtext

    cfg  = load_cfg()
    last = next(iter(cfg.values()), {}) if cfg else {}
    guess = last.get("last_libro") or find_workbook(".") or ""

    root = tk.Tk()
    root.title("Generador DIOT — SAT")
    root.resizable(False, False)
    frm = ttk.Frame(root, padding=14); frm.grid()

    ttk.Label(frm, text="Libro de registro (.xlsb / .xlsm / .xlsx):").grid(row=0, column=0, sticky="w")
    v_libro = tk.StringVar(value=guess)
    e = ttk.Entry(frm, textvariable=v_libro, width=64)
    e.grid(row=1, column=0, columnspan=2, sticky="we", pady=(0, 8))

    def browse():
        d = os.path.dirname(v_libro.get()) or "."
        p = filedialog.askopenfilename(
            title="Selecciona el Registro de Facturación",
            initialdir=d if os.path.isdir(d) else ".",
            filetypes=[("Libros de Excel (*.xlsb;*.xlsm;*.xlsx)", "*.xlsb *.xlsm *.xlsx"),
                       ("Libro Excel binario (*.xlsb)", "*.xlsb"),
                       ("Libro con macros (*.xlsm)", "*.xlsm"),
                       ("Libro Excel (*.xlsx)", "*.xlsx"),
                       ("Todos", "*.*")])
        if p: v_libro.set(p)

    ttk.Button(frm, text="Examinar…", command=browse).grid(row=1, column=2, padx=(8, 0), pady=(0, 8))

    ttk.Label(frm, text="Mes:").grid(row=2, column=0, sticky="w")
    v_mes = tk.StringVar(value=MES_LARGO[1])
    ttk.Combobox(frm, textvariable=v_mes, values=MES_LARGO[1:], state="readonly",
                 width=16).grid(row=3, column=0, sticky="w", pady=(0, 8))

    ttk.Label(frm, text="Tipo:").grid(row=2, column=1, sticky="w")
    v_tipo = tk.StringVar(value="N")
    ttk.Combobox(frm, textvariable=v_tipo,
                 values=["N", "C1", "C2", "C3", "C4", "C5"], width=8).grid(
                 row=3, column=1, sticky="w", pady=(0, 8))

    out = scrolledtext.ScrolledText(frm, width=84, height=18, font=("Consolas", 9))
    out.grid(row=5, column=0, columnspan=3, pady=(8, 0))

    def ask_path(name, initdir):
        return filedialog.asksaveasfilename(
            initialdir=initdir, initialfile=name, defaultextension=".txt",
            filetypes=[("Archivo DIOT (*.txt)", "*.txt")],
            title="Guardar DIOT como…") or None

    def go():
        libro = v_libro.get().strip()
        if not os.path.isfile(libro):
            messagebox.showerror("DIOT", "No encuentro el libro:\n%s" % libro); return
        month = MES_LARGO.index(v_mes.get())
        tipo  = v_tipo.get().strip().upper() or "N"
        if tipo != "N" and not re.fullmatch(r"C\d+", tipo):
            messagebox.showerror("DIOT", "Tipo inválido. Usa N o C1, C2, …"); return
        out.delete("1.0", "end"); out.insert("end", "Leyendo %s …\n" % libro); root.update()
        try:
            p, recs, tot, year, rfc, res = generate(libro, month, tipo, ask_path=ask_path)
        except Exception as ex:
            out.insert("end", "\nERROR: %s\n" % ex)
            messagebox.showerror("DIOT", str(ex)); return
        if not p:
            out.insert("end", "Cancelado.\n"); return
        _, lim, _ = read_control(libro)
        out.delete("1.0", "end")
        out.insert("end", summary(p, recs, tot, year, rfc, month, tipo, res,
                                  lim if lim is not None else LIMITE_EFE))

    ttk.Button(frm, text="Generar DIOT", command=go).grid(row=3, column=2, sticky="e", pady=(0, 8))
    root.mainloop()


# ---------- CLI ----------
def main():
    args = [a for a in sys.argv[1:]]
    libro = None
    if "--libro" in args:
        i = args.index("--libro"); libro = args[i + 1]; del args[i:i + 2]

    if not args or args[0] in ("--gui", "-g"):
        run_gui(); return

    month = int(args[0])
    tipo  = args[1].upper() if len(args) > 1 else "N"
    fixed = args[2] if len(args) > 2 else None
    if not 1 <= month <= 12:
        print("Mes inválido (1-12)."); sys.exit(1)
    if tipo != "N" and not re.fullmatch(r"C\d+", tipo):
        print("Tipo inválido. Usa N o C1, C2, …"); sys.exit(1)

    libro = libro or find_workbook(".")
    if not libro or not os.path.isfile(libro):
        print("No encontré el libro (.xlsb/.xlsm/.xlsx). "
              "Usa --libro \"ruta\\Registro.xlsm\" o corre sin argumentos para la GUI.")
        sys.exit(1)

    def ask_path(name, initdir):
        try:
            import tkinter as tk
            from tkinter import filedialog
            r = tk.Tk(); r.withdraw(); r.attributes("-topmost", True)
            p = filedialog.asksaveasfilename(
                initialdir=initdir, initialfile=name, defaultextension=".txt",
                filetypes=[("Archivo DIOT (*.txt)", "*.txt")], title="Guardar DIOT como…")
            r.destroy(); return p or None
        except Exception:
            return os.path.join(initdir, name)

    print("Leyendo %s ..." % libro)
    p, recs, tot, year, rfc, res = generate(libro, month, tipo, out_path=fixed, ask_path=ask_path)
    if not p:
        print("Cancelado por usuario."); sys.exit(0)
    _, lim, _ = read_control(libro)
    print(summary(p, recs, tot, year, rfc, month, tipo, res,
                  lim if lim is not None else LIMITE_EFE))


if __name__ == "__main__":
    main()
